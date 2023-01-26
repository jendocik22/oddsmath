import os
import sqlite3
import openpyxl

prj_dir = os.path.abspath(os.path.curdir)  # Получаем текущую папку проекта
a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
print(a)

odd_bf = sqlite3.connect('baza.db')
cursor = odd_bf.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS tab_1(oddsmath TEXT,betfair TEXT) ''')

file_to_read = openpyxl.load_workbook('baza.xlsx', data_only=True)
sheet = file_to_read['list']

for row in range(2, sheet.max_row + 1):  # Цикл по строкам начиная со второй (в первой заголовки)
    data = []  # Объявление списка
    for col in range(1, 3):  # Цикл по столбцам от 1 до 4 ( 5 не включая)
        value = sheet.cell(row, col).value  # value содержит значение ячейки с координатами row col
        data.append(value)  # Список который мы потом будем добавлять

    cursor.execute("INSERT INTO tab_1 VALUES (?, ?);", (data[0], data[1],))  # Вставка данных в поля таблицы
odd_bf.commit()  # сохраняем изменения
odd_bf.close()  # закрытие соединения
